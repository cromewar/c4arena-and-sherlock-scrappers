import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Load URLs from JSON
with open("c4competitions.json", "r") as file:
    urls = json.load(file)["urls"]

# Define the Excel file name
file_name = "scraped_data.xlsx"

# Initialize Selenium WebDriver (using Chrome)
driver = webdriver.Chrome()  # Ensure you have the correct WebDriver installed


def fetch_headlines(url):
    try:
        # Use Selenium to open the page
        driver.get(url)

        # Wait for the page to load completely
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "secondary-nav__item"))
        )

        # Initialize BeautifulSoup with the page source
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # Headlines title, dates, and prize pool
        headlines = soup.find_all(class_="type__headline__xs")

        # Leaderboard
        leaderboard = soup.find(class_="leaderboard-table-reduced__wrapper")

        # Data array
        data = []

        # Initialize a record dictionary for the combined data
        record = {}

        if headlines:
            for i in range(0, len(headlines), 5):
                # Extract registries
                name = headlines[i].get_text(strip=True)
                start_date = headlines[i + 1].get_text(strip=True)
                end_date = headlines[i + 2].get_text(strip=True)
                prize_pool = headlines[i + 3].get_text(strip=True)
                duration = headlines[i + 4].get_text(strip=True)

                # Print all the headlines found
                print(f"Name: {name}")
                print(f"Start date: {start_date}")
                print(f"End date: {end_date}")
                print(f"Prize pool: {prize_pool}")
                print(f"Duration: {duration}\n")

                # Store the headline information in the record
                record = {
                    "Name": name,
                    "Start date": start_date,
                    "End date": end_date,
                    "Prize pool": prize_pool,
                    "Duration": duration,
                    "Total": 0,
                    "High": 0,
                    "(Solo)": 0,
                    "Med": 0,
                    "Gas": 0,
                    "URL": url,  # Add the URL to the record
                }
        else:
            print("No headlines found")

        # If there is a leaderboard, switch to the "Details" tab
        if leaderboard:
            print("Leaderboard found")

            # Click on the "Details" tab
            details_tab = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//li[contains(text(),'Details')]")
                )
            )
            details_tab.click()

            # Allow some time for the content to load
            time.sleep(3)  # Adjust based on the page's load time

            # Reparse the page with the "Details" tab content
            soup = BeautifulSoup(driver.page_source, "html.parser")

            # Initialize a dictionary to hold the sum for each column
            sums = {}

            # Find all <ul> elements with role="row"
            rows = leaderboard.find_all("ul", role="row")

            # Iterate over rows starting from the second one (index 1)
            for row in rows[1:]:  # Skipping the first row which is the header
                items = row.find_all("li")

                # Start from the fourth item (index 3)
                for title, item in zip(rows[0].find_all("li")[3:], items[3:]):
                    text = item.get_text(strip=True)

                    # Try to convert the text to a float for summing
                    try:
                        value = float(text.replace(",", "").replace("$", ""))
                        column_name = title.get_text(strip=True)

                        # Add to the corresponding sum
                        if column_name in sums:
                            sums[column_name] += value
                        else:
                            sums[column_name] = value
                    except ValueError:
                        # If the conversion fails, skip this item (e.g., non-numeric values)
                        continue

            # Merge the sums dictionary into the record
            record.update(sums)

            # Print out the sums
            print("Sums:")
            for column_name, total_sum in sums.items():
                print(f"{column_name}: {total_sum}")
        else:
            print(
                "No leaderboard found with the class 'leaderboard-table-reduced__wrapper'."
            )

        # Find the nSloc data after switching tabs (or without switching if no leaderboard)
        table_container = soup.find_all("div", class_="table-container")

        for container in table_container:
            table = container.find("table")

            if table:
                rows = table.find_all("tr")

                for row in rows:
                    tds = row.find_all("td")
                    if any(
                        td.get_text(strip=True) in ["Totals", "TOTAL"] for td in tds
                    ):
                        # If "Totals" or "TOTAL" is found, handle accordingly
                        totals_data = [td.get_text(strip=True) for td in tds]

                        # Determine which position to extract based on the keyword and number of tds
                        if "Totals" in totals_data and len(totals_data) >= 4:
                            nsloc_value = totals_data[3]  # The fourth item (index 3)
                        elif "TOTAL" in totals_data and len(totals_data) >= 2:
                            nsloc_value = totals_data[1]  # The second item (index 1)

                        else:
                            print("Not enough data in the 'Totals'/'TOTAL' row.")
                            nsloc_value = "N/A"

                        print(f"nSloc: {nsloc_value}")

                        # Add nSloc to the record dictionary
                        record["nSloc"] = nsloc_value
                        break  # Stop searching after finding the relevant "Totals"/"TOTAL" row

        # Append the combined record to the data array
        data.append(record)

        # Convert the list of dictionaries to a DataFrame
        df = pd.DataFrame(data)

        # Append the data to Excel
        try:
            with pd.ExcelWriter(
                file_name, mode="a", engine="openpyxl", if_sheet_exists="overlay"
            ) as writer:
                df.to_excel(
                    writer,
                    index=False,
                    header=writer.sheets["Sheet1"].max_row == 0,
                    startrow=writer.sheets["Sheet1"].max_row,
                )
        except FileNotFoundError:
            df.to_excel(file_name, index=False)

    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")


# Example usage
for url in urls:
    fetch_headlines(url)

# Close the browser session after processing all URLs
driver.quit()
